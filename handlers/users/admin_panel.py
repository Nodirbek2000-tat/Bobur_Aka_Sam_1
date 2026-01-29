from aiogram import types
from aiogram.dispatcher import FSMContext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import tempfile
import json

from loader import dp, db
from keyboards.inline.buttons import get_admin_menu, get_stats_menu


async def is_admin(user_id: int) -> bool:
    return await db.is_admin(user_id)


@dp.message_handler(commands=['admin'], state='*')
async def cmd_admin(message: types.Message, state: FSMContext):
    await state.finish()

    if not await is_admin(message.from_user.id):
        await message.answer("⛔ Sizda ruxsat yo'q!")
        return

    total_users = await db.count_users()
    total_surveys = await db.count_surveys()
    total_channels = await db.count_channels()

    text = (
        "👨‍💼 <b>ADMIN PANEL</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{total_users}</b>\n"
        f"📋 So'rovnomalar: <b>{total_surveys}</b>\n"
        f"📢 Kanallar: <b>{total_channels}</b>\n\n"
        "Kerakli bo'limni tanlang:"
    )

    await message.answer(text, reply_markup=get_admin_menu())


@dp.callback_query_handler(text="admin:back", state='*')
async def callback_admin_back(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()

    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    total_users = await db.count_users()
    total_surveys = await db.count_surveys()
    total_channels = await db.count_channels()

    text = (
        "👨‍💼 <b>ADMIN PANEL</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{total_users}</b>\n"
        f"📋 So'rovnomalar: <b>{total_surveys}</b>\n"
        f"📢 Kanallar: <b>{total_channels}</b>\n\n"
        "Kerakli bo'limni tanlang:"
    )

    await callback.message.edit_text(text, reply_markup=get_admin_menu())
    await callback.answer()


@dp.callback_query_handler(text="admin:close", state='*')
async def callback_admin_close(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await callback.message.delete()
    await callback.answer()


@dp.callback_query_handler(text="admin:stats", state='*')
async def callback_stats(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    total_users = await db.count_users()
    users_24h = await db.count_users_last_24h()
    users_week = await db.count_users_last_week()

    active_survey = await db.get_active_survey()
    active_text = "Yo'q"
    responses_count = 0

    if active_survey:
        active_text = active_survey['name']
        responses_count = await db.count_survey_responses(active_survey['id'])

    text = (
        "📊 <b>STATISTIKA</b>\n\n"
        f"👥 <b>Foydalanuvchilar:</b>\n"
        f"   • Jami: <b>{total_users}</b>\n"
        f"   • Oxirgi 24 soat: <b>{users_24h}</b>\n"
        f"   • Oxirgi 7 kun: <b>{users_week}</b>\n\n"
        f"📋 <b>Aktiv so'rovnoma:</b> {active_text}\n"
        f"   • Javoblar soni: <b>{responses_count}</b>"
    )

    await callback.message.edit_text(text, reply_markup=get_stats_menu())
    await callback.answer()


@dp.callback_query_handler(text="stats:download", state='*')
async def callback_download_stats(callback: types.CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Sizda ruxsat yo'q!", show_alert=True)
        return

    active_survey = await db.get_active_survey()

    if not active_survey:
        await callback.answer("Aktiv so'rovnoma yo'q!", show_alert=True)
        return

    fields = await db.get_survey_fields(active_survey['id'])
    responses = await db.get_survey_responses(active_survey['id'])

    if not responses:
        await callback.answer("Javoblar yo'q!", show_alert=True)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Javoblar"

    headers = ["№"] + [f['column_name'] for f in fields] + ["Sana"]

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = border

    for row_num, response in enumerate(responses, 2):
        response_data = json.loads(response['response_data']) if isinstance(response['response_data'], str) else response['response_data']

        ws.cell(row=row_num, column=1, value=row_num-1).border = border

        for col, field in enumerate(fields, 2):
            value = response_data.get(field['column_name'], "")
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = border

        cell = ws.cell(row=row_num, column=len(headers), value=str(response['submitted_at']))
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    # 🔹 TO'G'RILANGAN JOY: temp papka Windows/Linux uchun
    file_path = os.path.join(tempfile.gettempdir(), active_survey['file_name'])
    wb.save(file_path)

    with open(file_path, 'rb') as file:
        await callback.message.answer_document(
            file,
            caption=f"📊 <b>{active_survey['name']}</b>\n👥 Jami javoblar: {len(responses)}"
        )

    os.remove(file_path)
    await callback.answer("Yuklandi!")
